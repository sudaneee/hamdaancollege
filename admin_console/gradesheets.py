"""
Excel gradesheet download/upload for exam result processing.

Flow: Exam Officer downloads a template for one Course + Session (the
roster comes from CourseRegistration, so only students actually registered
for that course can ever get a score) -> the filled sheet is handed to the
course lecturer outside the system -> the completed file is uploaded back
here, which upserts Result rows as drafts (or updates existing ones,
preserving whatever publish state they already had) -> the Exam Officer
reviews the roster on-screen and publishes when satisfied.
"""

from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from students.models import MAX_CA_SCORE, MAX_EXAM_SCORE, CourseRegistration, Result

HEADER_FILL = PatternFill(start_color='0B6B4B', end_color='0B6B4B', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def build_template(course, session, results_by_student):
    """Returns an openpyxl Workbook: one row per student registered for
    this course/session, pre-filled with whatever CA/Exam scores already
    exist so re-downloading for a correction doesn't lose prior entries."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Gradesheet'

    ws.append([f'{course.code} — {course.title}'])
    ws.append([f'Session: {session}  |  Semester: {course.semester}'])
    ws.append([f'CA is out of {MAX_CA_SCORE}, Exam is out of {MAX_EXAM_SCORE}. Do not edit the Student ID column.'])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(['Student ID', 'Name', 'CA', 'Exam'])
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    registrations = (
        CourseRegistration.objects.filter(course=course, session=session)
        .select_related('student').order_by('student__student_id')
    )
    for reg in registrations:
        student = reg.student
        existing = results_by_student.get(student.pk)
        ws.append([
            student.student_id, student.full_name,
            float(existing.ca_score) if existing else None,
            float(existing.exam_score) if existing else None,
        ])

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    return wb


class GradesheetImportResult:
    def __init__(self):
        self.created = 0
        self.updated = 0
        self.errors = []  # list of human-readable strings


def import_upload(course, session, file_obj):
    """Parses an uploaded gradesheet, upserting Result rows for students
    actually registered for this course/session. Invalid or unrecognised
    rows are skipped and reported, valid rows are always committed —
    partial success is better than an all-or-nothing reject on one typo."""
    outcome = GradesheetImportResult()

    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:
        outcome.errors.append(f'Could not read that file as an Excel spreadsheet ({exc}).')
        return outcome

    ws = wb.active

    roster = {
        reg.student.student_id: reg.student
        for reg in CourseRegistration.objects.filter(course=course, session=session).select_related('student')
    }

    # Find the header row (the one starting with "Student ID") rather than
    # assuming a fixed row number, so the template's title rows can't drift
    # out of sync with the parser.
    data_rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, row in enumerate(data_rows) if row and str(row[0]).strip().lower() == 'student id'),
        None,
    )
    if header_idx is None:
        outcome.errors.append('Could not find the "Student ID" header row — is this the downloaded template?')
        return outcome

    for row_num, row in enumerate(data_rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(cell in (None, '') for cell in row):
            continue
        student_id = str(row[0]).strip() if row[0] is not None else ''
        if not student_id:
            continue

        student = roster.get(student_id)
        if not student:
            outcome.errors.append(f'Row {row_num}: "{student_id}" is not registered for this course/session — skipped.')
            continue

        ca_raw, exam_raw = row[2], row[3]
        try:
            ca_score = Decimal(str(ca_raw)) if ca_raw not in (None, '') else Decimal('0')
            exam_score = Decimal(str(exam_raw)) if exam_raw not in (None, '') else Decimal('0')
        except InvalidOperation:
            outcome.errors.append(f'Row {row_num} ({student_id}): CA/Exam must be numbers — skipped.')
            continue

        if not (0 <= ca_score <= MAX_CA_SCORE):
            outcome.errors.append(f'Row {row_num} ({student_id}): CA must be between 0 and {MAX_CA_SCORE} — skipped.')
            continue
        if not (0 <= exam_score <= MAX_EXAM_SCORE):
            outcome.errors.append(f'Row {row_num} ({student_id}): Exam must be between 0 and {MAX_EXAM_SCORE} — skipped.')
            continue

        result, created = Result.objects.get_or_create(
            student=student, course=course, session=session,
            defaults={'semester': course.semester, 'ca_score': ca_score, 'exam_score': exam_score},
        )
        if not created:
            result.ca_score = ca_score
            result.exam_score = exam_score
        result.save()

        if created:
            outcome.created += 1
        else:
            outcome.updated += 1

    return outcome
